"""
excel_manager.py — SushiTegridad
Gestiona la lectura y escritura del archivo Excel que sirve como
base de datos única del sistema.

Uso:
    python excel_manager.py save   (lee datos de stdin, escribe Excel)
    python excel_manager.py load   (lee Excel, escribe datos a stdout)
"""
import sys
import os
from datetime import datetime, date

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] Se requiere openpyxl. Instale con: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

sys.stdout.reconfigure(encoding='utf-8')
sys.stdin.reconfigure(encoding='utf-8')

ARCHIVO_EXCEL = 'SushiTegridad_Analytics.xlsx'

# ═══════════════════════════════════════════════════════════════
#  ESTILOS PROFESIONALES
# ═══════════════════════════════════════════════════════════════
COLOR_HEADER   = PatternFill(start_color='1B1B2F', end_color='1B1B2F', fill_type='solid')
COLOR_ACCENT   = PatternFill(start_color='E43F5A', end_color='E43F5A', fill_type='solid')
COLOR_ROW_EVEN = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
COLOR_ALERT    = PatternFill(start_color='FFD6D6', end_color='FFD6D6', fill_type='solid')
COLOR_WARNING  = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
COLOR_OK       = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
COLOR_TOTAL    = PatternFill(start_color='2D2D44', end_color='2D2D44', fill_type='solid')

FONT_HEADER  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
FONT_TITLE   = Font(name='Calibri', bold=True, color='FFFFFF', size=14)
FONT_NORMAL  = Font(name='Calibri', size=10)
FONT_BOLD    = Font(name='Calibri', bold=True, size=10)
FONT_TOTAL   = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
FONT_ALERT   = Font(name='Calibri', bold=True, color='CC0000', size=10)
FONT_WARN    = Font(name='Calibri', bold=True, color='856404', size=10)
FONT_OK      = Font(name='Calibri', color='155724', size=10)

BORDER_THIN = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT   = Alignment(horizontal='left', vertical='center')
ALIGN_RIGHT  = Alignment(horizontal='right', vertical='center')


def escribir_titulo(ws, titulo, num_cols):
    """Escribe una fila de título con merge y estilo."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    celda = ws.cell(row=1, column=1, value=titulo)
    celda.font = FONT_TITLE
    celda.fill = COLOR_ACCENT
    celda.alignment = ALIGN_CENTER
    # Fecha de generación
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    celda2 = ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    celda2.font = Font(name='Calibri', italic=True, color='666666', size=9)
    celda2.alignment = ALIGN_CENTER


def escribir_headers(ws, fila, headers):
    """Escribe una fila de encabezados con estilo oscuro."""
    for col, texto in enumerate(headers, 1):
        celda = ws.cell(row=fila, column=col, value=texto)
        celda.font = FONT_HEADER
        celda.fill = COLOR_HEADER
        celda.alignment = ALIGN_CENTER
        celda.border = BORDER_THIN


def estilo_celda(celda, fila_par=False, alineacion=ALIGN_LEFT):
    """Aplica estilo base a una celda de datos."""
    celda.font = FONT_NORMAL
    celda.border = BORDER_THIN
    celda.alignment = alineacion
    if fila_par:
        celda.fill = COLOR_ROW_EVEN


def auto_ancho(ws):
    """Ajusta el ancho de cada columna al contenido."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


# ═══════════════════════════════════════════════════════════════
#  GUARDAR DATOS → EXCEL
# ═══════════════════════════════════════════════════════════════
def guardar():
    """Lee datos estructurados de stdin y genera el Excel profesional."""
    # Protección para no sobreescribir la plantilla avanzada (Dashboards)
    if os.path.exists(ARCHIVO_EXCEL):
        try:
            temp_wb = load_workbook(ARCHIVO_EXCEL, data_only=True)
            if 'Ventas Diarias' in temp_wb.sheetnames or 'Portada' in temp_wb.sheetnames:
                print(f"  [AVISO] Plantilla avanzada detectada en {ARCHIVO_EXCEL}. Se desactiva el guardado automático para proteger los gráficos y el formato.")
                return
        except Exception:
            pass

    lineas = sys.stdin.read().strip().split('\n')
    idx = 0

    def leer():
        nonlocal idx
        linea = lineas[idx].strip()
        idx += 1
        return linea

    # ── Parsear datos ─────────────────────────────────────────
    contrasena = leer()

    # Ingredientes
    num_ing = int(leer())
    ingredientes = []
    for _ in range(num_ing):
        partes = leer().split()
        ingredientes.append({
            'nombre': partes[0],
            'cantidad': float(partes[1]),
            'diasCaducidad': int(partes[2]),
            'fechaRegistro': partes[3]
        })

    # Platillos
    num_plat = int(leer())
    platillos = []
    for _ in range(num_plat):
        partes = leer().split()
        nombre = partes[0]
        precio = float(partes[1])
        num_receta = int(partes[2])
        receta = []
        for _ in range(num_receta):
            rp = leer().split()
            receta.append({'ingrediente': rp[0], 'cantidad': float(rp[1])})
        platillos.append({'nombre': nombre, 'precio': precio, 'receta': receta})

    # Ventas
    num_ventas = int(leer())
    ventas = []
    for _ in range(num_ventas):
        partes = leer().split()
        ventas.append({
            'platillo': partes[0],
            'cantidad': int(partes[1]),
            'total': float(partes[2])
        })

    # Empleados
    num_emp = int(leer())
    empleados = []
    for _ in range(num_emp):
        partes = leer().split()
        empleados.append({
            'nombre': partes[0],
            'puesto': partes[1],
            'sueldoDiario': float(partes[2]),
            'diasSemana': int(partes[3])
        })

    # ── Generar Excel ─────────────────────────────────────────
    wb = Workbook()

    # ─── Hoja: Inventario ─────────────────────────────────────
    ws_inv = wb.active
    ws_inv.title = 'Inventario'
    escribir_titulo(ws_inv, '📦 INVENTARIO — SUSHI TEGRIDAD', 6)
    headers_inv = ['ID', 'Ingrediente', 'Cantidad (Kg/Lt)', 'Caducidad (días)', 'Fecha Registro', 'Estado']
    escribir_headers(ws_inv, 4, headers_inv)

    for i, ing in enumerate(ingredientes):
        fila = 5 + i
        par = i % 2 == 0
        ws_inv.cell(row=fila, column=1, value=i + 1)
        ws_inv.cell(row=fila, column=2, value=ing['nombre'].replace('_', ' '))
        ws_inv.cell(row=fila, column=3, value=round(ing['cantidad'], 3))
        ws_inv.cell(row=fila, column=4, value=ing['diasCaducidad'] if ing['diasCaducidad'] > 0 else 'N/A')
        ws_inv.cell(row=fila, column=5, value=ing['fechaRegistro'])

        # Calcular estado
        if ing['diasCaducidad'] == 0:
            estado = 'No caduca'
            estado_fill = COLOR_OK
            estado_font = FONT_OK
        else:
            try:
                fecha_reg = date.fromisoformat(ing['fechaRegistro'])
                from datetime import timedelta
                fecha_cad = fecha_reg + timedelta(days=ing['diasCaducidad'])
                restantes = (fecha_cad - date.today()).days
                if restantes <= 0:
                    estado = '⚠ CADUCADO'
                    estado_fill = COLOR_ALERT
                    estado_font = FONT_ALERT
                elif restantes <= 2:
                    estado = f'⚠ Por caducar ({restantes}d)'
                    estado_fill = COLOR_WARNING
                    estado_font = FONT_WARN
                else:
                    estado = f'OK ({restantes}d)'
                    estado_fill = COLOR_OK
                    estado_font = FONT_OK
            except:
                estado = 'N/A'
                estado_fill = COLOR_ROW_EVEN
                estado_font = FONT_NORMAL

        celda_estado = ws_inv.cell(row=fila, column=6, value=estado)
        celda_estado.fill = estado_fill
        celda_estado.font = estado_font

        for col in range(1, 6):
            estilo_celda(ws_inv.cell(row=fila, column=col), par,
                         ALIGN_CENTER if col in (1, 4, 5) else (ALIGN_RIGHT if col == 3 else ALIGN_LEFT))
        ws_inv.cell(row=fila, column=6).alignment = ALIGN_CENTER
        ws_inv.cell(row=fila, column=6).border = BORDER_THIN

    auto_ancho(ws_inv)

    # ─── Hoja: Carta ──────────────────────────────────────────
    ws_carta = wb.create_sheet('Carta')
    escribir_titulo(ws_carta, '🍣 CARTA DE PLATILLOS — SUSHI TEGRIDAD', 4)
    headers_carta = ['No.', 'Platillo', 'Precio', 'Ingredientes (receta por unidad)']
    escribir_headers(ws_carta, 4, headers_carta)

    fila_carta = 5
    for i, plat in enumerate(platillos):
        par = i % 2 == 0
        receta_txt = ', '.join([f"{r['ingrediente'].replace('_',' ')} ({r['cantidad']:.3f})" for r in plat['receta']])
        ws_carta.cell(row=fila_carta, column=1, value=i + 1)
        ws_carta.cell(row=fila_carta, column=2, value=plat['nombre'].replace('_', ' '))
        celda_precio = ws_carta.cell(row=fila_carta, column=3, value=plat['precio'])
        celda_precio.number_format = '$#,##0.00'
        ws_carta.cell(row=fila_carta, column=4, value=receta_txt)
        for col in range(1, 5):
            estilo_celda(ws_carta.cell(row=fila_carta, column=col), par,
                         ALIGN_CENTER if col in (1, 3) else ALIGN_LEFT)
        fila_carta += 1

    auto_ancho(ws_carta)

    # ─── Hoja: Ventas ─────────────────────────────────────────
    ws_ventas = wb.create_sheet('Ventas')
    escribir_titulo(ws_ventas, '💰 HISTORIAL DE VENTAS — SUSHI TEGRIDAD', 4)
    headers_ventas = ['No.', 'Platillo', 'Cantidad', 'Total']
    escribir_headers(ws_ventas, 4, headers_ventas)

    total_ingresos = 0
    for i, v in enumerate(ventas):
        fila = 5 + i
        par = i % 2 == 0
        ws_ventas.cell(row=fila, column=1, value=i + 1)
        ws_ventas.cell(row=fila, column=2, value=v['platillo'].replace('_', ' '))
        ws_ventas.cell(row=fila, column=3, value=v['cantidad'])
        celda_total = ws_ventas.cell(row=fila, column=4, value=v['total'])
        celda_total.number_format = '$#,##0.00'
        total_ingresos += v['total']
        for col in range(1, 5):
            estilo_celda(ws_ventas.cell(row=fila, column=col), par,
                         ALIGN_CENTER if col in (1, 3) else (ALIGN_RIGHT if col == 4 else ALIGN_LEFT))

    # Fila de total
    if ventas:
        fila_total = 5 + len(ventas)
        ws_ventas.merge_cells(start_row=fila_total, start_column=1, end_row=fila_total, end_column=3)
        celda_label = ws_ventas.cell(row=fila_total, column=1, value='TOTAL INGRESOS')
        celda_label.font = FONT_TOTAL
        celda_label.fill = COLOR_TOTAL
        celda_label.alignment = ALIGN_RIGHT
        celda_label.border = BORDER_THIN
        for col in range(2, 4):
            ws_ventas.cell(row=fila_total, column=col).fill = COLOR_TOTAL
            ws_ventas.cell(row=fila_total, column=col).border = BORDER_THIN
        celda_ingreso = ws_ventas.cell(row=fila_total, column=4, value=total_ingresos)
        celda_ingreso.font = FONT_TOTAL
        celda_ingreso.fill = COLOR_TOTAL
        celda_ingreso.number_format = '$#,##0.00'
        celda_ingreso.alignment = ALIGN_RIGHT
        celda_ingreso.border = BORDER_THIN

    auto_ancho(ws_ventas)

    # ─── Hoja: Empleados ──────────────────────────────────────
    ws_emp = wb.create_sheet('Empleados')
    escribir_titulo(ws_emp, '👥 RECURSOS HUMANOS — SUSHI TEGRIDAD', 6)
    headers_emp = ['No.', 'Nombre', 'Puesto', 'Sueldo Diario', 'Días/Semana', 'Sueldo Semanal']
    escribir_headers(ws_emp, 4, headers_emp)

    for i, emp in enumerate(empleados):
        fila = 5 + i
        par = i % 2 == 0
        sueldo_semanal = emp['sueldoDiario'] * emp['diasSemana']
        ws_emp.cell(row=fila, column=1, value=i + 1)
        ws_emp.cell(row=fila, column=2, value=emp['nombre'].replace('_', ' '))
        ws_emp.cell(row=fila, column=3, value=emp['puesto'].replace('_', ' '))
        celda_sd = ws_emp.cell(row=fila, column=4, value=emp['sueldoDiario'])
        celda_sd.number_format = '$#,##0.00'
        ws_emp.cell(row=fila, column=5, value=emp['diasSemana'])
        celda_ss = ws_emp.cell(row=fila, column=6, value=sueldo_semanal)
        celda_ss.number_format = '$#,##0.00'
        for col in range(1, 7):
            estilo_celda(ws_emp.cell(row=fila, column=col), par,
                         ALIGN_CENTER if col in (1, 5) else (ALIGN_RIGHT if col in (4, 6) else ALIGN_LEFT))

    auto_ancho(ws_emp)

    # ─── Hoja oculta: _config (guarda contraseña) ─────────────
    ws_cfg = wb.create_sheet('_config')
    ws_cfg.cell(row=1, column=1, value='password')
    ws_cfg.cell(row=1, column=2, value=contrasena)
    ws_cfg.sheet_state = 'hidden'

    # ── Guardar ───────────────────────────────────────────────
    wb.save(ARCHIVO_EXCEL)
    print(f"  [OK] Excel guardado: {ARCHIVO_EXCEL}")


# ═══════════════════════════════════════════════════════════════
#  CARGAR DATOS ← EXCEL
# ═══════════════════════════════════════════════════════════════
def cargar():
    """Lee el Excel y escribe datos estructurados a stdout."""
    if not os.path.exists(ARCHIVO_EXCEL):
        print("NO_EXISTE")
        return

    try:
        wb = load_workbook(ARCHIVO_EXCEL, data_only=True)
    except Exception as e:
        print(f"ERROR:{e}")
        return

    # ── Contraseña ────────────────────────────────────────────
    if '_config' in wb.sheetnames:
        ws_cfg = wb['_config']
        contrasena = ws_cfg.cell(row=1, column=2).value or '1234'
    else:
        contrasena = '1234'
    print(contrasena)

    # ── Ingredientes ──────────────────────────────────────────
    if 'Inventario' in wb.sheetnames:
        ws = wb['Inventario']
        filas = []
        
        start_row = 5
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row[0] == 1 or str(row[0]) == '1':
                start_row = i
                break
                
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if row[0] is None:
                continue
            nombre = str(row[1]).replace(' ', '_') if row[1] else ''
            if not nombre: continue
            cantidad = float(row[2]) if row[2] else 0.0
            dias_cad_raw = row[3]
            if dias_cad_raw is None or str(dias_cad_raw).strip().upper() == 'N/A':
                dias_cad = 0
            else:
                try: dias_cad = int(dias_cad_raw)
                except ValueError: dias_cad = 0
            fecha = str(row[4]) if row[4] else date.today().isoformat()
            if ' ' in fecha:
                fecha = fecha.split(' ')[0]
            filas.append(f"{nombre} {cantidad:.6f} {dias_cad} {fecha}")
        print(len(filas))
        for f in filas:
            print(f)
    else:
        print(0)

    # ── Platillos ─────────────────────────────────────────────
    if 'Carta' in wb.sheetnames:
        ws = wb['Carta']
        platillos = []
        
        start_row = 5
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row[0] == 1 or str(row[0]) == '1':
                start_row = i
                break
                
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if row[0] is None:
                continue
            nombre = str(row[1]).replace(' ', '_') if row[1] else ''
            if not nombre: continue
            precio = float(row[2]) if row[2] else 0.0
            receta_txt = str(row[3]) if row[3] else ''
            ingredientes_receta = []
            if receta_txt and '(' in receta_txt and ')' in receta_txt:
                partes = receta_txt.split(', ')
                for parte in partes:
                    parte = parte.strip()
                    if '(' in parte and ')' in parte:
                        idx_paren = parte.rfind('(')
                        ing_nombre = parte[:idx_paren].strip().replace(' ', '_')
                        ing_cant = parte[idx_paren+1:parte.rfind(')')].strip()
                        try:
                            ingredientes_receta.append((ing_nombre, float(ing_cant)))
                        except ValueError:
                            pass
            platillos.append((nombre, precio, ingredientes_receta))

        print(len(platillos))
        for nombre, precio, receta in platillos:
            print(f"{nombre} {precio:.2f} {len(receta)}")
            for ing_n, ing_c in receta:
                print(f"{ing_n} {ing_c:.6f}")
    else:
        print(0)

    # ── Ventas ────────────────────────────────────────────────
    hoja_ventas_nombre = None
    for n in wb.sheetnames:
        if 'Ventas' in n:
            hoja_ventas_nombre = n
            break

    if hoja_ventas_nombre:
        ws = wb[hoja_ventas_nombre]
        ventas_filas = []
        is_advanced = ('Diarias' in hoja_ventas_nombre)
        
        start_row = 5
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row[0] == 1 or str(row[0]) == '1':
                start_row = i
                break

        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if row[0] is None:
                continue
            if isinstance(row[0], str) and ('TOTAL' in str(row[0]).upper() or 'NO.' in str(row[0]).upper()):
                break
                
            if is_advanced:
                platillo = str(row[5]).replace(' ', '_') if len(row) > 5 and row[5] else ''
                if not platillo: continue
                try:
                    cantidad = int(row[8]) if len(row) > 8 and row[8] else 0
                    total = float(row[9]) if len(row) > 9 and row[9] else 0.0
                except (ValueError, TypeError):
                    continue
                ventas_filas.append(f"{platillo} {cantidad} {total:.2f}")
            else:
                platillo = str(row[1]).replace(' ', '_') if row[1] else ''
                if not platillo: continue
                cantidad = int(row[2]) if row[2] else 0
                total = float(row[3]) if row[3] else 0.0
                ventas_filas.append(f"{platillo} {cantidad} {total:.2f}")
        print(len(ventas_filas))
        for f in ventas_filas:
            print(f)
    else:
        print(0)

    # ── Empleados ─────────────────────────────────────────────
    if 'Empleados' in wb.sheetnames:
        ws = wb['Empleados']
        emp_filas = []
        
        start_row = 5
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row[0] == 1 or str(row[0]) == '1':
                start_row = i
                break
                
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if row[0] is None:
                continue
            if isinstance(row[0], str) and 'TOTAL' in str(row[0]).upper():
                break
            nombre = str(row[1]).replace(' ', '_') if row[1] else ''
            if not nombre: continue
            puesto = str(row[2]).replace(' ', '_') if row[2] else ''
            try:
                sueldo = float(row[3]) if row[3] else 0.0
                dias = int(row[4]) if row[4] else 0
            except (ValueError, TypeError):
                sueldo = 0.0
                dias = 0
            emp_filas.append(f"{nombre} {puesto} {sueldo:.2f} {dias}")
        print(len(emp_filas))
        for f in emp_filas:
            print(f)
    else:
        print(0)


# ═══════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════
if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Uso: python excel_manager.py [save|load]", file=sys.stderr)
        sys.exit(1)

    modo = sys.argv[1].lower()
    if modo == 'save':
        guardar()
    elif modo == 'load':
        cargar()
    else:
        print(f"Modo desconocido: {modo}. Use 'save' o 'load'.", file=sys.stderr)
        sys.exit(1)
