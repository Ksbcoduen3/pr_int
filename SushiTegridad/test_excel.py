import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
wb = load_workbook('SushiTegridad_Analytics.xlsx', data_only=True)

for sheet in ['Inventario', 'Carta', 'Empleados']:
    if sheet in wb.sheetnames:
        print(f"\n--- {sheet} ---")
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True)):
            print(row)
